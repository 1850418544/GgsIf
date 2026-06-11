"""
Excel报告生成模块
使用openpyxl库生成格式化的Excel测试报告
支持添加测试用例、统计汇总、样式美化
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from config.config import Config
from utils.logger import logger

class ExcelReporter:
    """Excel报告生成器类"""
    
    def __init__(self):
        """初始化Excel报告
        
        创建工作簿和工作表，设置表头样式
        """
        # 创建新的工作簿
        self.wb = Workbook()
        # 获取活动工作表(默认第一个)
        self.ws = self.wb.active
        # 设置工作表名称
        self.ws.title = "测试报告"
        # 设置表头
        self.setup_header()
    
    def setup_header(self):
        """设置Excel表头
        
        定义表头列和样式，包括：
        - 用例编号、测试模块、用例名称、请求方法
        - 请求URL、请求参数、预期状态码、实际状态码
        - 实际响应、测试结果、响应时间
        - 错误信息、测试时间
        """
        # 表头列定义
        headers = [
            "用例编号", "测试模块", "用例名称", "请求方法", "请求URL",
            "请求参数", "预期状态码", "实际状态码", "实际响应",
            "测试结果", "响应时间(ms)", "错误信息", "测试时间"
        ]
        
        # 表头样式配置
        header_font = Font(bold=True, color="FFFFFF")  # 白色加粗字体
        header_fill = PatternFill(
            start_color="4472C4",  # 蓝色背景
            end_color="4472C4", 
            fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")  # 居中对齐
        
        # 单元格边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头并应用样式
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 设置列宽(根据内容调整)
        for col in range(1, len(headers) + 1):
            self.ws.column_dimensions[self.ws.cell(row=1, column=col).column_letter].width = 15
    
    def add_test_case(self, test_data):
        """添加测试用例数据到报告
        
        参数：
            test_data (dict): 测试用例数据，包含以下键：
                - case_id: 用例编号
                - test_module: 测试模块
                - case_name: 用例名称
                - method: 请求方法
                - url: 请求URL
                - params: 请求参数
                - expected_status: 预期状态码
                - actual_status: 实际状态码
                - actual_response: 实际响应
                - result: 测试结果(PASS/FAIL)
                - response_time: 响应时间
                - error_msg: 错误信息
                - test_time: 测试时间
        
        使用示例：
            reporter = ExcelReporter()
            reporter.add_test_case({
                "case_id": "001",
                "test_module": "登录模块",
                "case_name": "登录成功",
                "method": "POST",
                "result": "PASS"
            })
        """
        # 获取下一行行号
        row = self.ws.max_row + 1
        
        # 单元格边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 准备要写入的数据
        data = [
            test_data.get("case_id", ""),
            test_data.get("test_module", ""),
            test_data.get("case_name", ""),
            test_data.get("method", ""),
            test_data.get("url", ""),
            str(test_data.get("params", "")),
            test_data.get("expected_status", ""),
            test_data.get("actual_status", ""),
            str(test_data.get("actual_response", "")),
            test_data.get("result", ""),
            test_data.get("response_time", ""),
            test_data.get("error_msg", ""),
            test_data.get("test_time", "")
        ]
        
        # 写入数据并应用样式
        for col_num, value in enumerate(data, 1):
            cell = self.ws.cell(row=row, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)  # 自动换行
            
            # 测试结果列特殊处理(绿色通过，红色失败)
            if col_num == 10:
                if value == "PASS":
                    cell.font = Font(color="00B050")  # 绿色
                elif value == "FAIL":
                    cell.font = Font(color="FF0000")  # 红色
    
    def save_report(self, test_results):
        """保存Excel报告
        
        参数：
            test_results (list): 测试结果列表，每个元素是一个测试用例字典
        
        返回：
            str: 报告文件路径
        
        使用示例：
            reporter = ExcelReporter()
            for result in test_results:
                reporter.add_test_case(result)
            reporter.save_report(test_results)
        """
        # 确保目录存在
        Config.ensure_dirs()
        
        # 添加空行分隔内容
        self.ws.cell(row=self.ws.max_row + 1, column=1, value="")
        
        # 添加汇总信息
        summary_row = self.ws.max_row + 1
        self.ws.cell(row=summary_row, column=1, value="测试汇总")
        self.ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
        
        # 计算统计数据
        passed = sum(1 for r in test_results if r.get("result") == "PASS")
        failed = len(test_results) - passed
        
        # 写入统计信息
        self.ws.cell(row=summary_row + 1, column=1, value=f"总用例数: {len(test_results)}")
        self.ws.cell(row=summary_row + 2, column=1, value=f"通过: {passed}")
        self.ws.cell(row=summary_row + 3, column=1, value=f"失败: {failed}")
        self.ws.cell(row=summary_row + 4, column=1, value=f"通过率: {passed/len(test_results)*100:.2f}%" if test_results else "0%")
        self.ws.cell(row=summary_row + 5, column=1, value=f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # 保存报告文件
        try:
            # 添加时间戳避免文件冲突
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_path = Config.REPORT_DIR / f"test_report_{timestamp}.xlsx"
            self.wb.save(report_path)
            logger.info(f"Excel报告已生成: {report_path}")
            return report_path
        except Exception as e:
            logger.error(f"生成Excel报告失败: {e}")
            raise